from django.http import HttpResponse
from django.db.models import Sum, Q
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Company, LedgerEntry


def calculate_ledger_data(company, start_date=None, end_date=None):
    """Calculate ledger data for a company with date range"""
    entries = LedgerEntry.objects.filter(company=company).order_by(
        'transaction_date', 'created_at'
    )

    # Filter by date range if provided
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            entries = entries.filter(transaction_date__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            entries = entries.filter(transaction_date__lte=end_date_obj)
        except ValueError:
            pass

    # Calculate opening balance
    opening_balance = Decimal('0.00')
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            opening_entries = LedgerEntry.objects.filter(
                company=company,
                transaction_date__lt=start_date_obj
            )
            opening_debit = opening_entries.filter(transaction_type='debit').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            opening_credit = opening_entries.filter(transaction_type='credit').aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            opening_balance = opening_debit - opening_credit
        except ValueError:
            pass

    # Calculate running balance
    running_balance = opening_balance
    entries_list = []

    for entry in entries:
        if entry.transaction_type == 'debit':
            running_balance += entry.amount
            debit_amount = entry.amount
            credit_amount = None
        else:
            running_balance -= entry.amount
            debit_amount = None
            credit_amount = entry.amount

        entries_list.append({
            'date': entry.transaction_date,
            'reference': entry.transaction_number,
            'description': entry.description or '',
            'debit': debit_amount,
            'credit': credit_amount,
            'balance': running_balance,
            'transaction_type': entry.transaction_type
        })

    # Calculate totals
    total_debit = entries.filter(transaction_type='debit').aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    total_credit = entries.filter(transaction_type='credit').aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    closing_balance = opening_balance + total_debit - total_credit

    return {
        'company': company,
        'opening_balance': opening_balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': closing_balance,
        'entries': entries_list,
        'start_date': start_date,
        'end_date': end_date
    }


def export_ledger_pdf(company, start_date=None, end_date=None):
    """Export company ledger as PDF"""
    data = calculate_ledger_data(company, start_date, end_date)

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#283593'),
        spaceAfter=12
    )
    normal_style = styles['Normal']

    # Title
    elements.append(Paragraph("COMPANY LEDGER", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Company Information
    company_info = [
        ['Company Name:', data['company'].name],
        ['Email:', data['company'].email or 'N/A'],
        ['Phone:', data['company'].phone or 'N/A'],
        ['Address:', data['company'].address or 'N/A'],
    ]

    if data['start_date'] or data['end_date']:
        date_range = []
        if data['start_date']:
            date_range.append(f"From: {data['start_date']}")
        if data['end_date']:
            date_range.append(f"To: {data['end_date']}")
        company_info.append(['Date Range:', ' - '.join(date_range)])

    company_table = Table(company_info, colWidths=[2 * inch, 4 * inch])
    company_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Summary
    summary_data = [
        ['Opening Balance', f"₹ {data['opening_balance']:,.2f}"],
        ['Total Debit', f"₹ {data['total_debit']:,.2f}"],
        ['Total Credit', f"₹ {data['total_credit']:,.2f}"],
        ['Closing Balance', f"₹ {data['closing_balance']:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3949ab')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#5c6bc0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.white),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Ledger Entries
    if data['entries']:
        elements.append(Paragraph("Transaction Details", heading_style))
        
        # Table header
        table_data = [['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']]
        
        # Add entries
        for entry in data['entries']:
            debit_str = f"₹ {entry['debit']:,.2f}" if entry['debit'] else ""
            credit_str = f"₹ {entry['credit']:,.2f}" if entry['credit'] else ""
            table_data.append([
                entry['date'].strftime('%Y-%m-%d'),
                entry['reference'],
                entry['description'][:50] if entry['description'] else '',  # Truncate long descriptions
                debit_str,
                credit_str,
                f"₹ {entry['balance']:,.2f}"
            ])

        # Create table
        ledger_table = Table(table_data, colWidths=[0.8 * inch, 1.2 * inch, 2 * inch, 1 * inch, 1 * inch, 1 * inch])
        ledger_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Right align amounts
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(ledger_table)
    else:
        elements.append(Paragraph("No transactions found for the selected period.", normal_style))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    # Create HTTP response
    filename = f"ledger_{company.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_ledger_excel(company, start_date=None, end_date=None):
    """Export company ledger as Excel"""
    data = calculate_ledger_data(company, start_date, end_date)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Ledger"

    # Styles
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    title_fill = PatternFill(start_color="3949ab", end_color="3949ab", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    normal_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "COMPANY LEDGER"
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Company Information
    row = 3
    ws[f'A{row}'] = 'Company Name:'
    ws[f'B{row}'] = data['company'].name
    row += 1
    ws[f'A{row}'] = 'Email:'
    ws[f'B{row}'] = data['company'].email or 'N/A'
    row += 1
    ws[f'A{row}'] = 'Phone:'
    ws[f'B{row}'] = data['company'].phone or 'N/A'
    row += 1
    ws[f'A{row}'] = 'Address:'
    ws[f'B{row}'] = data['company'].address or 'N/A'
    
    if data['start_date'] or data['end_date']:
        row += 1
        date_range = []
        if data['start_date']:
            date_range.append(f"From: {data['start_date']}")
        if data['end_date']:
            date_range.append(f"To: {data['end_date']}")
        ws[f'A{row}'] = 'Date Range:'
        ws[f'B{row}'] = ' - '.join(date_range)

    # Format company info
    for r in range(3, row + 1):
        ws[f'A{r}'].font = Font(bold=True)
        ws[f'A{r}'].fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")

    row += 2

    # Summary
    ws[f'A{row}'] = 'Opening Balance'
    ws[f'B{row}'] = f"₹ {data['opening_balance']:,.2f}"
    row += 1
    ws[f'A{row}'] = 'Total Debit'
    ws[f'B{row}'] = f"₹ {data['total_debit']:,.2f}"
    row += 1
    ws[f'A{row}'] = 'Total Credit'
    ws[f'B{row}'] = f"₹ {data['total_credit']:,.2f}"
    row += 1
    ws[f'A{row}'] = 'Closing Balance'
    ws[f'B{row}'] = f"₹ {data['closing_balance']:,.2f}"
    
    # Format summary
    for r in range(row - 3, row + 1):
        ws[f'A{r}'].fill = header_fill
        ws[f'A{r}'].font = header_font
        ws[f'B{r}'].fill = header_fill
        ws[f'A{r}'].alignment = left_align
        ws[f'B{r}'].alignment = right_align
        ws[f'B{r}'].font = header_font

    row += 2

    # Ledger Entries Header
    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    row += 1

    # Add entries
    for entry in data['entries']:
        ws.cell(row=row, column=1, value=entry['date'].strftime('%Y-%m-%d')).border = border
        ws.cell(row=row, column=2, value=entry['reference']).border = border
        ws.cell(row=row, column=3, value=entry['description']).border = border
        ws.cell(row=row, column=4, value=f"₹ {entry['debit']:,.2f}" if entry['debit'] else "").border = border
        ws.cell(row=row, column=5, value=f"₹ {entry['credit']:,.2f}" if entry['credit'] else "").border = border
        ws.cell(row=row, column=6, value=f"₹ {entry['balance']:,.2f}").border = border
        
        # Right align amount columns
        ws.cell(row=row, column=4).alignment = right_align
        ws.cell(row=row, column=5).alignment = right_align
        ws.cell(row=row, column=6).alignment = right_align
        
        # Alternate row colors
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
        
        row += 1

    # Auto-adjust column widths
    column_widths = [12, 18, 40, 15, 15, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create HTTP response
    filename = f"ledger_{company.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
